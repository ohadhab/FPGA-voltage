import serial
import time
from datetime import datetime
from openpyxl import Workbook
from openpyxl.utils.cell import get_column_letter
import keyboard  # using module keyboard

##########################################
# !DO NOT OPEN EXCEL FILE WHILE RUNNING! #
# log.txt file can be open WHILE RUNNING #
##########################################
# Stop the program with CTRL+F1 and wait #
#           to next iteration            #
##########################################

time_between_lines = 1  # seconds between lines the FPGA sends to UART
time_to_run = 0  # seconds to run, 0-inf
UART_COM = 6  # the communication port that the FPGA is connected to
Baud_Rate = 115200  # the Baud Rate that the FPGA is transmitting

stop = False
stop2 = False


def ccttrrll(event):
    global stop
    if event.event_type == "down":
        stop = True
        # print('ctrl')
    else:
        stop = False
        # print('clean ctrl')


def ff11(event):
    global stop
    global stop2
    if event.event_type == "down" and stop:
        stop2 = True
        print('Stopping')
    else:
        stop = False
        # print('clean ctrl')


# ---------> hook event handler
keyboard.hook_key('ctrl', ccttrrll)
keyboard.hook_key('f1', ff11)
# --------->

reg = 11  # number of regs
COM = 'COM' + str(UART_COM)
itr = 0
ser = serial.Serial(COM, Baud_Rate, timeout=10 * time_between_lines)  # timeout- prevent stuck while wait new line
begin = datetime.now()
begin = begin.strftime("%d-%m-%Y %H-%M-%S.%f") + '.xlsx'
wb = Workbook()
wb.save(begin)
sheet = wb.active
sheet['A' + str(1)].value = 'Date and Time'

# for x in range(reg):
#     sheet[get_column_letter(x + 2) + str(1)].value = 'reg no.' + str(1 + x)
#     sheet[get_column_letter(x + 2 + reg) + str(1)].value = 'actual reg no.' + str(1 + x)
sheet[get_column_letter(2) + str(1)].value = 'Temp hex'
sheet[get_column_letter(3) + str(1)].value = 'Vp-Vn(0-1) hex'
sheet[get_column_letter(4) + str(1)].value = 'A0(0-3.3) hex'
sheet[get_column_letter(5) + str(1)].value = 'A1(0-3.3) hex'
sheet[get_column_letter(6) + str(1)].value = 'A2(0-3.3) hex'
sheet[get_column_letter(7) + str(1)].value = 'A3(0-3.3) hex'
sheet[get_column_letter(8) + str(1)].value = 'A4(0-3.3) hex'
sheet[get_column_letter(9) + str(1)].value = 'A5(0-3.3) hex'
sheet[get_column_letter(10) + str(1)].value = 'A6-A7(0-1) hex'
sheet[get_column_letter(11) + str(1)].value = 'A8-A9(0-1) hex'
sheet[get_column_letter(12) + str(1)].value = 'A10-A11(0-1) hex'

sheet[get_column_letter(13) + str(1)].value = 'Temp'
sheet[get_column_letter(14) + str(1)].value = 'Vp-Vn(0-1)'
sheet[get_column_letter(15) + str(1)].value = 'A0(0-3.3)'
sheet[get_column_letter(16) + str(1)].value = 'A1(0-3.3)'
sheet[get_column_letter(17) + str(1)].value = 'A2(0-3.3)'
sheet[get_column_letter(18) + str(1)].value = 'A3(0-3.3)'
sheet[get_column_letter(19) + str(1)].value = 'A4(0-3.3)'
sheet[get_column_letter(20) + str(1)].value = 'A5(0-3.3)'
sheet[get_column_letter(21) + str(1)].value = 'A6-A7(0-1)'
sheet[get_column_letter(22) + str(1)].value = 'A8-A9(0-1)'
sheet[get_column_letter(23) + str(1)].value = 'A10-A11(0-1)'

wb.save(begin)
# opening log.txt file
f = open('log.txt', 'a')
f.write("\n" + begin + ", time between lines: " + str(time_between_lines) + "s, time to run: " + str(time_to_run) +
        "s, number of regs: " + str(reg) + ", Baud Rate: " + str(Baud_Rate) + ", UART " + COM + "\n")
f.close()
print((begin + ", time between lines: " + str(time_between_lines) + "s, time to run: " + str(time_to_run) +
       "s, number of regs: " + str(reg) + ", Baud Rate: " + str(Baud_Rate) + ", UART " + COM))

while time_to_run == 0 or (itr - 1) * time_between_lines <= time_to_run:
    data = ser.readline()
    print(data)
    if itr > 0:
        data = str(data)
        data = data[4:-3]
        data_splitted = data.split()
        # datetime object containing current date and time
        now = datetime.now()
        # dd/mm/YY H:M:S.us
        now = now.strftime("%d/%m/%Y %H:%M:%S.%f")

        sheet['A' + str(itr + 1)].value = now
        for x in range(reg):
            sheet[get_column_letter(x + 2) + str(itr + 1)].value = data_splitted[x]

        sheet[get_column_letter(13) + str(itr + 1)].value = '=HEX2DEC(B'+str(itr + 1)+')*503.975/4096-273.15'
        sheet[get_column_letter(14) + str(itr + 1)].value = '=HEX2DEC(C'+str(itr + 1)+')/4096'
        sheet[get_column_letter(15) + str(itr + 1)].value = '=(HEX2DEC(D'+str(itr + 1)+')/4096*3.3)-((-0.013159)*(HEX2DEC(D'+str(itr + 1)+')/4096*3.3)+0.0075857)'
        sheet[get_column_letter(16) + str(itr + 1)].value = '=(HEX2DEC(E'+str(itr + 1)+')/4096*3.3)-((-0.013159)*(HEX2DEC(E'+str(itr + 1)+')/4096*3.3)+0.0075857)'
        sheet[get_column_letter(17) + str(itr + 1)].value = '=(HEX2DEC(F'+str(itr + 1)+')/4096*3.3)-((-0.013159)*(HEX2DEC(F'+str(itr + 1)+')/4096*3.3)+0.0075857)'
        sheet[get_column_letter(18) + str(itr + 1)].value = '=(HEX2DEC(G'+str(itr + 1)+')/4096*3.3)-((-0.013159)*(HEX2DEC(G'+str(itr + 1)+')/4096*3.3)+0.0075857)'
        sheet[get_column_letter(19) + str(itr + 1)].value = '=(HEX2DEC(H'+str(itr + 1)+')/4096*3.3)-((-0.013159)*(HEX2DEC(H'+str(itr + 1)+')/4096*3.3)+0.0075857)'
        sheet[get_column_letter(20) + str(itr + 1)].value = '=(HEX2DEC(I'+str(itr + 1)+')/4096*3.3)-((-0.013159)*(HEX2DEC(I'+str(itr + 1)+')/4096*3.3)+0.0075857)'
        sheet[get_column_letter(21) + str(itr + 1)].value = '=HEX2DEC(J'+str(itr + 1)+')/4096'
        sheet[get_column_letter(22) + str(itr + 1)].value = '=HEX2DEC(K'+str(itr + 1)+')/4096'
        sheet[get_column_letter(23) + str(itr + 1)].value = '=HEX2DEC(L'+str(itr + 1)+')/4096'

        wb.save(begin)
        print(str(itr) + " " + now + " " + data)
        # opening log.txt file
        f = open('log.txt', 'a')
        f.write(str(itr) + " " + now + " " + data + "\n")
        f.close()
    if stop2:
        print('Stopped by user')
        f = open('log.txt', 'a')
        f.write("Stopped by user\n")
        f.close()
        break
    itr = itr + 1
    time.sleep(0.5 * time_between_lines)  # time.sleep- ease on the computer

print('DONE')
# opening log.txt file
f = open('log.txt', 'a')
f.write("DONE\n")
f.close()
exit(0)
